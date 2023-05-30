import openpyxl
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker 
import datetime
from openpyxl import Workbook
import openpyxl

# Global Variable(s) and Function(s)
path_240    = './../excel_files/filtered_ITET_data_240[km].xlsx'
path_275    = './../excel_files/filtered_ITET_data.xlsx'

name_of_file = "data_for_armeet"

# Excell Specifics #######################################
# data = ['Time', 'TI_at_275', 'Error_TI', 'ne_at_275', 'Error_ne', 'File_name']
data = ['Time', 'TI_at_275', 'TE', 'TI_name_of_file']

ans = input("Would you like to save this data into a file [Y/n]? ")

if (ans == 'y' or ans == 'Y'):
    try:
        wb = openpyxl.load_workbook('./../excel_files/' + name_of_file + '.xlsx')
        ws = wb.active
        print("Found file", end="\n\n")
    except:
        wb = Workbook()
        ws = wb.active
        ws.append(data)
        print("Create the file", end='\n\n')
     
     
def EPOCH_to_DATE(epoch_time):
    date_conv = datetime.datetime.fromtimestamp(epoch_time)
    return date_conv

def subtract_time(hours, minutes):
    time_in_minutes = hours*60 + minutes

    time_in_minutes -= (7*60+45)

    if (time_in_minutes < 0):
        time_in_minutes += 24*60
    
    hours   = time_in_minutes // 60
    minutes = time_in_minutes % 60

    return hours, minutes
###################################################

# Open the workbook 
wb_obj_240 = openpyxl.load_workbook(path_240)
sheet_obj_240 = wb_obj_240.active

wb_obj_275 = openpyxl.load_workbook(path_275)
sheet_obj_275 = wb_obj_275.active

y = []
x = []

for i in range(1, sheet_obj_275.max_row):
    
    date_275    = sheet_obj_275.cell(row=i+1, column=1).value
    y_slot_275  = sheet_obj_275.cell(row=i+1, column=2).value 

    for j in range(1, sheet_obj_275.max_row):
        # date_240    = EPOCH_to_DATE(sheet_obj_240.cell(row=i+1, column=1).value) 
        date_240    = sheet_obj_240.cell(row=j+1, column=1).value 
        y_slot_240  = sheet_obj_240.cell(row=j+1, column=2).value 
    
    
        if (y_slot_240 <= 20000 and 
            y_slot_275 <= 20000 and
            abs(date_240 - date_275) <= (30 * 60) 
            ):
            
            date_at_275    = EPOCH_to_DATE(sheet_obj_275.cell(row=i+1, column=1).value) 
            year    = date_at_275.year
            time    = date_at_275.time()
  
            hours, minutes = subtract_time(time.hour, time.minute)
            x.append(hours + minutes / 60.0)
            y.append(y_slot_275)
    
            TE                      = sheet_obj_275.cell(row=i+1, column=3).value
            name_of_IT_file       = sheet_obj_275.cell(row=i+1, column=4).value

            # ['Time', 'TI_at_275', 'TE', 'TI_name_of_file']
            if (ans == 'y' or ans == 'Y'):
                try: 
                    ws.append( ( 
                        date_275,
                        y_slot_275,
                        TE,
                        name_of_IT_file
                    ) )
                except:
                    print("Error-------------------------------")
                    print("Smth went wrong when saving the file")
                    print("------------------------------------")

if (ans == 'y' or ans == 'Y'):
    wb.save('./../excel_files/' + name_of_file + '.xlsx')

pre_noon        = 0
post_midnight   = 6
post_noon       = 12
pre_midnight    = 18
night           = 24 

day_time = [ post_midnight, pre_noon, post_noon, pre_midnight, night ]

fig = plt.figure(figsize=(14, 10))

# Bins for the histograms
bins = []
i = 0 
while(i <= 24):
    bins.append(i)
    i += 3 

# Scatter plot
ax1 = fig.add_subplot(211)
ax1.scatter(x, y, color="orange", label="IT: " + str(len(y)))
ax1.grid(True)
ax1.set_title('Ti at 275 km - filtered with 240[km] data')
ax1.legend()

# Histogram plot #1
ax2 = fig.add_subplot(212)
ax2.hist(x, bins=bins, edgecolor='black')
for i in day_time: 
    ax1.axvline(x = i, color='r')
    ax2.axvline(x = i, color='r')

plt.plot()
plt.show()