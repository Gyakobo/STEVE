import openpyxl
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker 
import datetime
from openpyxl import Workbook
import openpyxl
import os

from unicodedata import name
import h5py as h5
import matplotlib.pyplot as plt
import datetime
import time

# Global Variable(s) and Function(s)
dir_path    = "/home/andrew/PFISR_data/"
# path        = '/home/andrew/STEVE/New_API/excel_files/Dataset_2014-2016.xlsx'
path        = '/home/andrew/STEVE/New_API/excel_files/Dataset_2014-2016_240[km].xlsx'

def get_files(path):
    for file in os.listdir(path):
        if os.path.isfile(os.path.join(path, file)):
            yield file

### GENERATOR EXPRESSION ###
# for file in get_files(dir_path):
#    print(file)

def EPOCH_to_DATE(epoch_time):
    date_conv = datetime.datetime.fromtimestamp(epoch_time)
    return date_conv

def subtract_time(hours, minutes):
    time_in_minutes = hours*60 + minutes

    time_in_minutes -= (7*60+45)

    if (time_in_minutes < 0):
        time_in_minutes += 24*60
    
    hours   = time_in_minutes // 60
    minutes = time_in_minutes % 60

    return hours, minutes
###################################################

# Open the workbook 
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

y = []
x = []
name_of_file = []
n = []

### GENERATOR EXPRESSION ###
for file in get_files(dir_path):
    # print(dir_path+file, '###################')
    
    try:
        f  = h5.File(dir_path+file, 'r')
    except:
        print("Cannot open file")

    beamcodes = f['BeamCodes']
    beamcodesdata = beamcodes[:,:]
    beamcodesdata = beamcodesdata.astype(float)    
    
    # EPOCH TIME
    epoch = f['UnixTime']
    epochData = epoch[:,:]
    epochData = epochData.astype(float)    

    # CONVERTS EPOCH TIME TO UNIVERSAL TIME IN DATETIME 
    '''
    ut = [datetime.datetime(int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[0]), # YEAR
                           int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[1]), # MONTH
                           int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[2]), # DAY
                           int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[3]), # HOUR
                           int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[4]), # MINUTE
                           int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[5])) # SECOND
                           for t in range(len(epochData))]
    '''
    
    for t in range(len(epochData)):
        h = int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[3]) # HOUR
        m = int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[4]) # MINUTE
        h, m = subtract_time(h, m)
        n.append(h + m / 60.0)


for i in range(1, sheet_obj.max_row): # sheet_obj.max_row
    
    date    = EPOCH_to_DATE(sheet_obj.cell(row=i+1, column=1).value) 
    y_slot  = sheet_obj.cell(row=i+1, column=2).value 

    file_name_explicit  = sheet_obj.cell(row=i+1, column=6).value
    time_explicit       = str(sheet_obj.cell(row=i+1, column=1).value)

    name_of_file.append(file_name_explicit + ', ' + time_explicit)
    # TIspike20150529.002_lp_5min.h5

    if (y_slot <= 20000):        
        year    = date.year
        time    = date.time()

        hours, minutes = subtract_time(time.hour, time.minute)
        x.append(hours + minutes / 60.0)
        y.append(y_slot)


pre_noon        = 0
post_midnight   = 6
post_noon       = 12
pre_midnight    = 18
night           = 24 

day_time = [ post_midnight, pre_noon, post_noon, pre_midnight, night ]

fig = plt.figure(figsize=(14, 10))

# Bins for the histograms
n_bin_del = 0.5 
bins = []
i = 0 
while(i <= 24):
    bins.append(i)
    i += n_bin_del

# Scatter plot
# ax1 = fig.add_subplot(311)
ax1 = fig.add_subplot(211)
ax1.scatter(x, y, color="orange", label="IT: " + str(len(y)))
ax1.grid(True)
ax1.set_title('Ti at 240 km')
ax1.legend()

# Histogram plot #1
ax2 = fig.add_subplot(212)
# ax2 = fig.add_subplot(312)
ax2.hist(x, bins=bins, edgecolor='black')

# Histogram plot #2
# ax3 = fig.add_subplot(313)
# ax3.hist(n, bins=bins, edgecolor='black')

# Trace the hour red lines
for i in day_time: 
    ax1.axvline(x = i, color='r')
    # ax2.axvline(x = i, color='r')
    # ax3.axvline(x = i, color='r')

plt.plot()
plt.show()

############################### Deprecated ##################################

# plt.scatter(x, y, color="orange", label="IT: " + str(len(y)))
# plt.grid(True)
# plt.title('Ti at 275 km')
# plt.subplots_adjust(bottom=0.162)

# for i in day_time: 
#    plt.axvline(x = i, color='r')

# plt.xlabel("IT: " + str(sheet_obj.max_row) + " points found")
# plt.ylabel('Line-of-Sight Ion Temperature [K]')
# plt.legend()
