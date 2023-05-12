import openpyxl
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker 
import datetime
from openpyxl import Workbook
import openpyxl

# Global Variable(s) and Function(s)
path    = '/home/andrew/STEVE/New_API/excel_files/filtered_ITET_data.xlsx'

def EPOCH_to_DATE(epoch_time):
    date_conv = datetime.datetime.fromtimestamp(epoch_time)
    return date_conv

def subtract_time(hours, minutes):
    time_in_minutes = hours*60 + minutes

    time_in_minutes -= (7*60+45)

    if (time_in_minutes < 0):
        time_in_minutes += 24*60
    
    hours   = time_in_minutes // 60
    minutes = time_in_minutes % 60

    return hours, minutes
###################################################

# Open the workbook 
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

y = []
x = []
name_of_file = []

for i in range(1, sheet_obj.max_row): # sheet_obj.max_row
    
    date    = EPOCH_to_DATE(sheet_obj.cell(row=i+1, column=1).value) 
    y_slot  = sheet_obj.cell(row=i+1, column=2).value 
    
    if y_slot >= 5000:
        name_of_file.append(sheet_obj.cell(row=i+1, column=4).value[7:-3]) 
    else: name_of_file.append(None) 

    if (y_slot <= 20000):        
        year    = date.year 
        time    = date.time()
 
        hours, minutes = subtract_time(time.hour, time.minute)
        x.append(hours + minutes / 60.0)
        y.append(y_slot)


pre_noon        = 0
post_midnight   = 6
post_noon       = 12
pre_midnight    = 18
night           = 24 

day_time = [ post_midnight, pre_noon, post_noon, pre_midnight, night ]

fig = plt.figure(figsize=(14, 10))

# Bins for the histograms
bins = []
i = 0 
while(i <= 24):
    bins.append(i)
    i += 0.5 

# Scatter plot
# ax1 = fig.add_subplot(211)
ax1 = fig.add_subplot(111)

ax1.scatter(x, y, color="orange", label="IT: " + str(len(y)))

for i, txt in enumerate(name_of_file):
    ax1.annotate(txt, (x[i], y[i]))

ax1.grid(True)
ax1.set_title('Ti at 275 km')
ax1.legend()

# Histogram plot #1
# ax2 = fig.add_subplot(212)
# ax2.hist(x, bins=bins, edgecolor='black')
for i in day_time: 
    ax1.axvline(x = i, color='r')
    # ax2.axvline(x = i, color='r')

plt.plot()
plt.show()

############################### Deprecated ##################################

# plt.scatter(x, y, color="orange", label="IT: " + str(len(y)))
# plt.grid(True)
# plt.title('Ti at 275 km')
# plt.subplots_adjust(bottom=0.162)

# for i in day_time: 
#    plt.axvline(x = i, color='r')

# plt.xlabel("IT: " + str(sheet_obj.max_row) + " points found")
# plt.ylabel('Line-of-Sight Ion Temperature [K]')
# plt.legend()