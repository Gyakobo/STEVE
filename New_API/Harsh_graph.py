import openpyxl
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker 
import datetime
import geopy.distance
import os
from openpyxl import Workbook
import openpyxl
import colorama
import sys
import numpy as np

# Global Variable(s) and Function(s)
path = '/home/andrew/STEVE/New_API/excel_files/Dataset_2014-2016.xlsx'
# path = '/home/andrew/STEVE/New_API/excel_files/Dataset_2015.xlsx'

harsh_path_swarm_a = '/home/andrew/STEVE/New_API/SWARM_Data/Harsh_files_SWARM_A/swarm_A_2014-2017.xlsx'
harsh_path_swarm_b = '/home/andrew/STEVE/New_API/SWARM_Data/Harsh_files_SWARM_B/swarm_B_2014-2017.xlsx'
harsh_path_swarm_c = '/home/andrew/STEVE/New_API/SWARM_Data/Harsh_files_SWARM_C/swarm_C_2014-2017.xlsx'

poker_flat_lat  = 65.1200
poker_flat_long = -147.4700 

def scatter_hist(x, y, ax, ax_histx, ax_histy):
    # no labels
    ax_histx.tick_params(axis="x", labelbottom=False)
    ax_histy.tick_params(axis="y", labelleft=False)

    # the scatter plot:
    ax.scatter(x, y)

    # now determine nice limits by hand:
    binwidth = 0.25
    xymax = max(np.max(np.abs(x)), np.max(np.abs(y)))
    lim = (int(xymax/binwidth) + 1) * binwidth

    bins = np.arange(-lim, lim + binwidth, binwidth)
    ax_histx.hist(x, bins=bins)
    ax_histy.hist(y, bins=bins, orientation='horizontal')   


def progress_bar(progress, total, color=colorama.Fore.YELLOW):
    percent = 100 * (progress / float(total))
    bar = '█' * int(percent) + '-' * (100-int(percent))
    print(color + f"\r|{bar}| {percent:.2f}%", end="\r")
    # sys.stdout.write(color + f"\r|{bar}| {percent:.2f}%")
    # sys.stdout.flush()

def curv_distance_km(lat1, long1, lat2, long2):
    coords_1 = (lat1, long1)
    coords_2 = (lat2, long2)
    return geopy.distance.geodesic(coords_1, coords_2).km

def curv_distance_miles(lat1, long1, lat2, long2):
    coords_1 = (lat1, long1)
    coords_2 = (lat2, long2)
    return geopy.distance.geodesic(coords_1, coords_2).miles

def EPOCH_to_DATE(epoch_time):
    date_conv = datetime.datetime.fromtimestamp(epoch_time)
    #return date_conv.strftime("%d-%m, %H:%M:%S")
    return date_conv

def EPOCH_to_min(epoch_time):
    date_conv = datetime.datetime.fromtimestamp(epoch_time)
    return int(date_conv.strftime("%M"))

def subtract_time(hours, minutes):
    time_in_minutes = hours*60 + minutes

    time_in_minutes -= (7*60+45)

    if (time_in_minutes < 0): time_in_minutes += 24*60
    
    hours   = time_in_minutes // 60
    minutes = time_in_minutes % 60
    
    return hours, minutes
#################################


# Open the workbook 
wb_obj = openpyxl.load_workbook(path, read_only=True)
sheet_obj = wb_obj.active

wb_obj_swarm_a = openpyxl.load_workbook(harsh_path_swarm_a)
sheet_obj_swarm_a = wb_obj_swarm_a.active
wb_obj_swarm_b = openpyxl.load_workbook(harsh_path_swarm_b)
sheet_obj_swarm_b = wb_obj_swarm_b.active
wb_obj_swarm_c = openpyxl.load_workbook(harsh_path_swarm_c)
sheet_obj_swarm_c = wb_obj_swarm_c.active

y_a = []
x_a = []

y_b = []
x_b = []

y_c = []
x_c = []

x = []
y = []

# for i in range(1, sheet_obj.max_row): # sheet_obj.max_row
for i in range(1, 10): # sheet_obj.max_row
    progress_bar(i+1, sheet_obj.max_row)

    x_slot = sheet_obj.cell(row=i+1, column=1).value 
    y_slot = sheet_obj.cell(row=i+1, column=2).value 

    for j in range(1, sheet_obj_swarm_a.max_row):
        harsh_epoch_time =  sheet_obj_swarm_a.cell(row=j+1, column=1).value.timestamp()
        harsh_te    =       sheet_obj_swarm_a.cell(row=j+1, column=2).value
        harsh_long  =       sheet_obj_swarm_a.cell(row=j+1, column=3).value
        harsh_lat   =       sheet_obj_swarm_a.cell(row=j+1, column=4).value
        
        if abs(int(x_slot) - int(harsh_epoch_time)) <= (30.0 * 60.0) and curv_distance_km(harsh_lat, harsh_long, poker_flat_lat, poker_flat_long) <= 500:
            date    = sheet_obj_swarm_a.cell(row=j+1, column=1).value
            year    = date.year #sheet_obj_swarm_a.cell(row=j+1, column=1).value
            time    = date.time()

            hours, minutes = subtract_time(time.hour, time.minute)
            alaska_datetime = datetime.datetime(
                    year=2000, month=1, day=1, hour=hours, minute=minutes,
                    second=time.second, microsecond=time.microsecond) # tzinfo=pytz.utc

            if (y_slot <= 20000):        
                x_a.append(alaska_datetime)
                y_a.append(y_slot)
    
    for j in range(1, sheet_obj_swarm_b.max_row):
        harsh_epoch_time =  sheet_obj_swarm_b.cell(row=j+1, column=1).value.timestamp()
        harsh_te    =       sheet_obj_swarm_b.cell(row=j+1, column=2).value
        harsh_long  =       sheet_obj_swarm_b.cell(row=j+1, column=3).value
        harsh_lat   =       sheet_obj_swarm_b.cell(row=j+1, column=4).value
        if abs(int(x_slot) - int(harsh_epoch_time)) <= (30.0 * 60.0) and curv_distance_km(harsh_lat, harsh_long, poker_flat_lat, poker_flat_long) <= 500:
            date    = sheet_obj_swarm_b.cell(row=j+1, column=1).value
            year    = date.year #sheet_obj_swarm_a.cell(row=j+1, column=1).value
            time    = date.time()

            hours, minutes = subtract_time(time.hour, time.minute)
            alaska_datetime = datetime.datetime(
                    year=2000, month=1, day=1, hour=hours, minute=minutes,
                    second=time.second, microsecond=time.microsecond) # tzinfo=pytz.utc

            if (y_slot <= 20000):        
                x_b.append(alaska_datetime)
                y_b.append(y_slot)

    for j in range(1, sheet_obj_swarm_c.max_row):
        harsh_epoch_time =  sheet_obj_swarm_c.cell(row=j+1, column=1).value.timestamp()
        harsh_te    =       sheet_obj_swarm_c.cell(row=j+1, column=2).value
        harsh_long  =       sheet_obj_swarm_c.cell(row=j+1, column=3).value
        harsh_lat   =       sheet_obj_swarm_c.cell(row=j+1, column=4).value
        if abs(int(x_slot) - int(harsh_epoch_time)) <= (30.0 * 60.0) and curv_distance_km(harsh_lat, harsh_long, poker_flat_lat, poker_flat_long) <= 500:
            date    = sheet_obj_swarm_c.cell(row=j+1, column=1).value
            year    = date.year #sheet_obj_swarm_a.cell(row=j+1, column=1).value
            time    = date.time()

            hours, minutes = subtract_time(time.hour, time.minute)
            alaska_datetime = datetime.datetime(
                    year=2000, month=1, day=1, hour=hours, minute=minutes,
                    second=time.second, microsecond=time.microsecond) # tzinfo=pytz.utc

            if (y_slot <= 20000):        
                x_c.append(alaska_datetime)
                y_c.append(y_slot)
         
                '''try: 
                    ws.append( ( 
                        harsh_epoch_time, x_slot,
                        harsh_lat, harsh_long, 
                        harsh_te, y_slot 
                    ) )
                except:
                    print("Smth went wrong")'''

# wb.save('./harsh_andy_excel_files/' + name_of_file + '.xlsx')

IT_storms = ( 
[1476139380, 1476475320],
[1457058240, 1457620560],
[1462641000, 1462978080],
[1453203960, 1453483860],
[1471662420, 1472106180],
[1477335480, 1478023800]
)

post_midnight   = datetime.datetime(2000, 1, 1, 0, 0, 0)
pre_noon        = datetime.datetime(2000, 1, 1, 6, 0, 0)
post_noon       = datetime.datetime(2000, 1, 1, 12, 0, 0)
pre_midnight    = datetime.datetime(2000, 1, 1, 18, 0, 0)
day_time = [ post_midnight, pre_noon, post_noon, pre_midnight ]

# Default console to default color
print(colorama.Fore.RESET)

fig = plt.figure(figsize=(14, 9))

# A, B, C
plt.scatter(x_a, y_a, color="orange", label="SWARM A: " + str(len(y_a)))
plt.scatter(x_b, y_b, color="green", label="SWARM B: " + str(len(y_b)))
plt.scatter(x_c, y_c, color="blue", label="SWARM C: " + str(len(y_c)))
plt.xticks(rotation='vertical')

plt.grid(True)

plt.title('Ti at 275 km')
plt.subplots_adjust(bottom=0.162)

for i in day_time: 
    plt.axvline(x = i, color='r')

for i in IT_storms: 
    plt.axvline(x = EPOCH_to_DATE(i[0]), color='r')
    plt.axvline(x = EPOCH_to_DATE(i[1]), color='g')

plt.xlabel("IT - ET, " + str(sheet_obj.max_row) + " points; " + 
str(len(y_a) + len(y_b) + len(y_c)) + " found")

# plt.xlabel('Epoch Time(Day-Month, Hour:Minute:Second)')
plt.ylabel('Line-of-Sight Ion Temperature [K]')

plt.legend()
plt.plot()
plt.show()


