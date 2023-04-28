import openpyxl
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker 
import datetime
import geopy.distance
import os
from openpyxl import Workbook
import openpyxl
import colorama
import sys
import numpy as np


name_of_file = "filtered_ITET_data"

# Excell Specifics #######################################
# data = ['Time', 'TI_at_275', 'Error_TI', 'ne_at_275', 'Error_ne', 'File_name']
data = ['Time', 'TI_at_275', 'TE', 'TI_name_of_file']

try:
    wb = openpyxl.load_workbook('./excel_files/' + name_of_file + '.xlsx')
    ws = wb.active

    print("Found file", end="\n\n")

except:
    wb = Workbook()
    ws = wb.active
    ws.append(data)

    print("Create the file", end='\n\n')



# Global Variable(s) and Function(s)
path = '/home/andrew/STEVE/New_API/excel_files/Dataset_2014-2016.xlsx'
# path = '/home/andrew/STEVE/New_API/excel_files/Dataset_2015.xlsx'

harsh_path_swarm_a = '/home/andrew/STEVE/New_API/SWARM_Data/Harsh_files_SWARM_A/swarm_A_2014-2017.xlsx'
harsh_path_swarm_b = '/home/andrew/STEVE/New_API/SWARM_Data/Harsh_files_SWARM_B/swarm_B_2014-2017.xlsx'
harsh_path_swarm_c = '/home/andrew/STEVE/New_API/SWARM_Data/Harsh_files_SWARM_C/swarm_C_2014-2017.xlsx'

poker_flat_lat  = 65.1200
poker_flat_long = -147.4700 

def progress_bar(progress, total, color=colorama.Fore.YELLOW):
    percent = 100 * (progress / float(total))
    bar = '█' * int(percent) + '-' * (100-int(percent))
    print(color + f"\r|{bar}| {percent:.2f}%", end="\r")
    # sys.stdout.write(color + f"\r|{bar}| {percent:.2f}%")
    # sys.stdout.flush()

def curv_distance_km(lat1, long1, lat2, long2):
    coords_1 = (lat1, long1)
    coords_2 = (lat2, long2)
    return geopy.distance.geodesic(coords_1, coords_2).km

def curv_distance_miles(lat1, long1, lat2, long2):
    coords_1 = (lat1, long1)
    coords_2 = (lat2, long2)
    return geopy.distance.geodesic(coords_1, coords_2).miles

def EPOCH_to_DATE(epoch_time):
    date_conv = datetime.datetime.fromtimestamp(epoch_time)
    #return date_conv.strftime("%d-%m, %H:%M:%S")
    return date_conv

def EPOCH_to_min(epoch_time):
    date_conv = datetime.datetime.fromtimestamp(epoch_time)
    return int(date_conv.strftime("%M"))

def subtract_time(hours, minutes):
    time_in_minutes = hours*60 + minutes

    time_in_minutes -= (7*60+45)

    if (time_in_minutes < 0): time_in_minutes += 24*60
    
    hours   = time_in_minutes // 60
    minutes = time_in_minutes % 60
    
    return hours, minutes
#################################


# Open the workbook 
wb_obj = openpyxl.load_workbook(path, read_only=True)
sheet_obj = wb_obj.active

wb_obj_swarm_a = openpyxl.load_workbook(harsh_path_swarm_a)
sheet_obj_swarm_a = wb_obj_swarm_a.active
wb_obj_swarm_b = openpyxl.load_workbook(harsh_path_swarm_b)
sheet_obj_swarm_b = wb_obj_swarm_b.active
wb_obj_swarm_c = openpyxl.load_workbook(harsh_path_swarm_c)
sheet_obj_swarm_c = wb_obj_swarm_c.active

y_a = []
x_a = []

y_b = []
x_b = []

y_c = []
x_c = []

x = []
y = []

n = sheet_obj.max_row
# n = 30
for i in range(1, n): # sheet_obj.max_row
    progress_bar(i+1, n)

    x_slot          = sheet_obj.cell(row=i+1, column=1).value 
    y_slot          = sheet_obj.cell(row=i+1, column=2).value 
    name_of_IT_file = sheet_obj.cell(row=i+1, column=6).value

    for j in range(1, sheet_obj_swarm_a.max_row):
        harsh_epoch_time =  sheet_obj_swarm_a.cell(row=j+1, column=1).value.timestamp()
        harsh_te    =       sheet_obj_swarm_a.cell(row=j+1, column=2).value
        harsh_long  =       sheet_obj_swarm_a.cell(row=j+1, column=3).value
        harsh_lat   =       sheet_obj_swarm_a.cell(row=j+1, column=4).value
        
        if abs(int(x_slot) - int(harsh_epoch_time)) <= (30.0 * 60.0) and curv_distance_km(harsh_lat, harsh_long, poker_flat_lat, poker_flat_long) <= 500:
            date    = sheet_obj_swarm_a.cell(row=j+1, column=1).value
            # year    = date.year
            time    = date.time()

            hours, minutes = subtract_time(time.hour, time.minute)

            if (y_slot <= 20000):        
                x_a.append(hours + minutes / 60.0)
                y_a.append(y_slot)
                
                try: 
                    ws.append( ( 
                        x_slot,
                        y_slot,
                        harsh_te,
                        name_of_IT_file
                    ) )
                except:
                    print("Smth went wrong")
    
    for j in range(1, sheet_obj_swarm_b.max_row):
        harsh_epoch_time =  sheet_obj_swarm_b.cell(row=j+1, column=1).value.timestamp()
        harsh_te    =       sheet_obj_swarm_b.cell(row=j+1, column=2).value
        harsh_long  =       sheet_obj_swarm_b.cell(row=j+1, column=3).value
        harsh_lat   =       sheet_obj_swarm_b.cell(row=j+1, column=4).value
        if abs(int(x_slot) - int(harsh_epoch_time)) <= (30.0 * 60.0) and curv_distance_km(harsh_lat, harsh_long, poker_flat_lat, poker_flat_long) <= 500:
            date    = sheet_obj_swarm_b.cell(row=j+1, column=1).value
            # year    = date.year 
            time    = date.time()

            hours, minutes = subtract_time(time.hour, time.minute)

            if (y_slot <= 20000):        
                x_b.append(hours + minutes / 60.0)
                y_b.append(y_slot)
                
                try: 
                    ws.append( ( 
                        x_slot,
                        y_slot,
                        harsh_te,
                        name_of_IT_file
                    ) )
                except:
                    print("Smth went wrong")

    for j in range(1, sheet_obj_swarm_c.max_row):
        harsh_epoch_time =  sheet_obj_swarm_c.cell(row=j+1, column=1).value.timestamp()
        harsh_te    =       sheet_obj_swarm_c.cell(row=j+1, column=2).value
        harsh_long  =       sheet_obj_swarm_c.cell(row=j+1, column=3).value
        harsh_lat   =       sheet_obj_swarm_c.cell(row=j+1, column=4).value
        if abs(int(x_slot) - int(harsh_epoch_time)) <= (30.0 * 60.0) and curv_distance_km(harsh_lat, harsh_long, poker_flat_lat, poker_flat_long) <= 500:
            date    = sheet_obj_swarm_c.cell(row=j+1, column=1).value
            # year    = date.year
            time    = date.time()

            hours, minutes = subtract_time(time.hour, time.minute)
            
            if (y_slot <= 20000):        
                x_c.append(hours + minutes / 60.0)
                y_c.append(y_slot)
                
                try: 
                    ws.append( ( 
                        x_slot,
                        y_slot,
                        harsh_te,
                        name_of_IT_file
                    ) )
                except:
                    print("Smth went wrong")

wb.save('./excel_files/' + name_of_file + '.xlsx')

# Default console to default color
print(colorama.Fore.RESET)

fig = plt.figure(figsize=(14, 9))

# plt.xlabel("IT - ET, " + str(sheet_obj.max_row) + " points; " + 
# str(len(y_a) + len(y_b) + len(y_c)) + " found")

pre_noon        = 0
post_midnight   = 6
post_noon       = 12
pre_midnight    = 18
night           = 24 

day_time = [ post_midnight, pre_noon, post_noon, pre_midnight, night ]

fig = plt.figure(figsize=(14, 10))

x.append(x_a)
x.append(x_b)
x.append(x_c)

'''
y.append(y_a)
y.append(y_b)
y.append(y_c)
'''

# Bins for the histograms
bins = []
i = 0 
while(i <= 24):
    bins.append(i)
    i += 0.5 

# Scatter plot
ax1 = fig.add_subplot(211)

# A, B, C
ax1.scatter(x_a, y_a, color="orange", label="SWARM A: " + str(len(y_a)))
ax1.scatter(x_b, y_b, color="green", label="SWARM B: " + str(len(y_b)))
ax1.scatter(x_c, y_c, color="blue", label="SWARM C: " + str(len(y_c)))
ax1.grid(True)
ax1.set_title('Ti at 275 km')
ax1.legend()

# Histogram plot #1
ax2 = fig.add_subplot(212)
ax2.hist(x, bins=bins, edgecolor='black')
for i in day_time: 
    ax1.axvline(x = i, color='r')
    ax2.axvline(x = i, color='r')

plt.plot()
plt.show()

