# THIS CODE WILL PRODUCE ION TEMPERATURE FIGURES FOR LONG-PULSE DATAFILES 
# DOWNLOADED FROM "data.amisr.com" AND THEN STRIPPED USING "Code-for-Andy-1.py"


# IMPORT THE REQUIRED LIBRARIES (DO NOT TOUCH, UNLESS ALTERING THE CODE)
from random import sample
from unicodedata import name
import h5py as h5
import numpy as np
import matplotlib.pyplot as plt
import math
import datetime
import time
import sys

from openpyxl import Workbook
import openpyxl


name_of_file = "Dataset_2018"


# Excell Specifics #######################################
data = ['Time', 'TI_at_275', 'Error_TI', 'ne_at_275', 'Error_ne', 'File_name']
try:
  wb = openpyxl.load_workbook('./excel_files/' + name_of_file + '.xlsx')
  ws = wb.active

  print("Found file", end="\n\n")

except:
  wb = Workbook()
  ws = wb.active
  ws.append(data)

  print("Create the file", end='\n\n')

  #cell_obj = ws.cell(row=1, column=1)
  #if cell_obj != "Time":
  #  ws.append(data)

# TODO list
# Time - make a normal date [year, month, date]
# ne - Plasma density
###########################################################







# My Edits ################################################
from numpy import NaN, Inf, arange, isscalar, asarray, array
from math import *
# EVIL Peak finding function
from endolith_peakdet import *




# General data file structure


# Check whether the argument is a NaN
def isNaN(num):
  if float('-inf') < float(num) < float('inf'):
    return False
  else:
    return True

# Gets the peaks and the given plunges of a given array
# Dr.Gareth asked to change the THRESHOLD
def get_peaks(data, ne_data, Ti_error, ne_error, epochData, THRESHOLD = 2):
  m_data = []
  m_ne_data = []
  m_ti_error = []
  m_ne_error = []
  m_list_epochData = [] 

  print("\n\n\n")
  print("Sample variable #######")
  print(epochData[0])
  print("#######################\n\n\n")

  #####################
  # Golden Number = 15 #
  #####################

  for idx, item in enumerate(data):
    if not isNaN(item[0]):
      m_data.append(item[0])
      m_ne_data.append(ne_data[idx])
      m_ti_error.append(Ti_error[idx])
      m_ne_error.append(ne_error[idx])
      m_list_epochData.append(epochData.tolist()[idx])
  
  #data = m_data
  #ne_data = m_ne_data

  print('len of data: ' + str(len((m_data))))
  print('len of ne_data: ' + str(len((m_ne_data))))
  print('len of epochData: ' + str(len(m_list_epochData)), end='\n\n')

  mean    = sum(m_data) / len(m_data)
  std_dev = (sum( [ (val-mean)**2 for val in m_data ] )/ (len(m_data)-1))**0.5
  peaks, plunges = peakdet(m_data, THRESHOLD*std_dev)

  peaks_x = []
  plunges_x = []

  ne_at_275_lst = []
  ti_error_lst = []
  ne_error_lst = []

  #peaks_x   = [ (val, ) for val in array(peaks)[:,0] ] #Peaks per index
  peaks_y   = [ (val, ) for val in array(peaks)[:,1] ]
  for var in peaks_y:
    index = m_data.index(var[0])
    peaks_x.append(m_list_epochData[index]) 
    ne_at_275_lst.append(m_ne_data[index])
    ti_error_lst.append(m_ti_error[index])
    ne_error_lst.append(m_ne_error[index])

  print("Working on file" + sys.argv[1] + "------------------------------")
  print('THESE ARE PEAKS ON THE X-AXIS')
  print(np.array(peaks_x))
  print('------------------------------------------')
  print('THESE ARE PEAKS ON THE Y-AXIS')
  print(np.array(peaks_y))
  print("----------------------------------------------------------------", end="\n\n")
  
  #plunges_x   = [ (val, ) for val in array(plunges)[:,0] ] # Plunges per index
  plunges_y   = [ (val, ) for val in array(plunges)[:,1] ]
  for var in plunges_y:
    index = m_data.index(var[0])
    plunges_x.append(m_list_epochData[index]) 




  '''
  # Save to excel file col A --------------------------------------------
  for index, peak in enumerate(peaks_x): 
    ws['A' + str(index+2)] = peak[0]
  # ---------------------------------------------------------------------
  
  # Save to excel file col B --------------------------------------------
  for index, peak in enumerate(peaks_y): 
    ws['B' + str(index+2)] = peak[0]
  # ---------------------------------------------------------------------
  
  # Save to excel file col C --------------------------------------------
  for index, val in enumerate(ti_error_lst): 
    ws['C' + str(index+2)] = val[0]
  # ---------------------------------------------------------------------

  # Save to excel file col D --------------------------------------------
  for index, val in enumerate(ne_at_275_lst): 
    ws['D' + str(index+2)] = val[0]
  # ---------------------------------------------------------------------
  
  # Save to excel file col E and col F ----------------------------------
  for index, val in enumerate(ne_error_lst): 
    ws['E' + str(index+2)] = val[0]
    ws['F' + str(index+2)] = sys.argv[1]
  # ---------------------------------------------------------------------
  '''

  for index, val in enumerate(peaks_x): 
    ws.append( (val[0], peaks_y[index][0], ti_error_lst[index][0], ne_at_275_lst[index][0], ne_error_lst[index][0], sys.argv[1]) )

  # Save the file
  #wb.save('./excel_files/' + sys.argv[1] + '.xlsx')
  wb.save('./excel_files/' + name_of_file + '.xlsx')

  return np.array(peaks_x), np.array(peaks_y), np.array(plunges_x), np.array(plunges_y), std_dev, mean

###########################################################








# CUSTOMIZE THE PATH TO THE DATAFILE RELATIVE TO THE CODE
path = "/home/andrew/PFISR_data/"

# INSERT THE NAME OF THE h5 DATAFILE YOU HAVE DOWNLOADED AND WOULD LIKE TO PLOT
# filename = "TIspike20220406.005_lp_1min-fitcal.h5"
filename = sys.argv[1]

# THIS OPENS THE DATA FILE OF INTEREST
try:
  f  = h5.File(path+filename, 'r')
except:
  print("Cannot open file")

beamcodes = f['BeamCodes']
beamcodesdata = beamcodes[:,:]
beamcodesdata = beamcodesdata.astype(float)    

# EPOCH TIME
epoch = f['UnixTime']
epochData = epoch[:,:]
epochData = epochData.astype(float)    

# ALTITUDE
alt = f['Altitude']
altdata = alt[:,:]/1000.#CONVERTING FROM METERS TO KILOMETERS
altdata = altdata.astype(float)    

# PLASMA DENSITY [m-3]
ne = f['Ne']
nedata = ne[:,:,:]
nedata = nedata.astype(float)    

# ERROR IN PLASMA DENSITY [m-3]
dne = f['dNe']
dnedata = dne[:,:,:]
dnedata = dnedata.astype(float)    

# THE ION SPECIES IS SET
# 0 = O+, THE DOMINANT F-REGION ION. SOME DATAFILES OFFER, 1 = O2+, 2 = NO+ , 
# 3 = N2+ AND , 4 = N+, BUT I STRONGLY SUGGEST YOU DO NOT TOUCH THIS
ion = 0

# ION TEMPERATURE [K]
ti = f['Fits']
tidata = ti[:,:,:,ion,1]
tidata = tidata.astype(float)    

# ERROR IN ION TEMPERATURE [K]
dti = f['Errors']
dtidata = dti[:,:,:,ion,1]
dtidata = dtidata.astype(float)    

# ELECTRON TEMPERATURE [K]
te = f['Fits']
tedata = te[:,:,:,-1,1]  
tedata = tedata.astype(float)    

# ERROR IN ELECTRON TEMPERATURE [K]
dte = f['Errors']
dtedata = dte[:,:,:,-1,1]
dtedata = dtedata.astype(float)    

# LINE-OF-SIGHT ION VELOCITY [m/s]
losv = f['Fits']
losvdata = losv[:,:,:,ion,3]
losvdata = losvdata.astype(float)    

# CORRECTED GEOMAGNETIC LATITUDE [deg]
lat = f['MagneticLatitude']
latdata = lat[:,:]
latdata = latdata.astype(float)       

# CORRECTED GEOMAGNETIC LONGITUDE [deg]
long = f['MagneticLongitude']
longdata = long[:,:]
longdata = longdata.astype(float)  

# THIS CLOSES THE DATA FILE
f.close() 

# THIS CONVERTS EPOCH TIME TO UNIVERSAL TIME IN DATETIME
ut = [datetime.datetime(int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[0]), # YEAR
                       int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[1]), # MONTH
                       int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[2]), # DAY
                       int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[3]), # HOUR
                       int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[4]), # MINUTE
                       int(time.gmtime((epochData[t,0]+epochData[t,1])/2)[5])) # SECOND
                       for t in range(len(epochData))]

# ERRONEOUS POINTS ARE REMOVED (COMMENT THIS OUT TO VIEW ALL MEASUREMENTS)

for tt in range(len(nedata)):
  for ii in range(len(nedata[0])):
    for jj in range(len(nedata[0,0])):
      if nedata[tt,ii,jj] < dnedata[tt,ii,jj]:
        nedata[tt,ii,jj] = np.NaN
      if tidata[tt,ii,jj] < dtidata[tt,ii,jj]:
        tidata[tt,ii,jj] = np.NaN
      if tedata[tt,ii,jj] < dtedata[tt,ii,jj]:
        tedata[tt,ii,jj] = np.NaN

# WE CYCLE THROUGH THE DIFFERENT BEAMS IN THE EXPERIMENT
for i in range(len(altdata)):
# WE FOCUS ON THE BEAM PARALLEL TO THE MAGNETIC FIELD (BEAM CODE 64157)    
  if beamcodesdata[i,0] == 64157.0:

    #WE FIND THE ION TEMPERATURE MEASURED CLOSEST TO 275 KM      
    a = abs(altdata[i,:] - 275)
    b = np.where(a == np.nanmin(a))
    Ti275     = tidata[:,i,b[0]]
    ne275     = nedata[:,i,b[0]]
    Ti_error  = dtidata[:,i,b[0]]
    ne_error  = dnedata[:,i,b[0]]




    # Still my edits ##################################
    try:
      # Threshold = 3.5
      peaks_x, peaks_y, plunges_x, plunges_y, std_dev, mean = get_peaks( 
              Ti275.tolist(), 
              ne275.tolist(), 
              Ti_error.tolist(), 
              ne_error.tolist(),
              epochData, 
              THRESHOLD=3.5 
            )
      #epochData = get_peaks( Ti275.tolist(), epochData, THRESHOLD=3.5 )     
    except:
      print('Something went wrong')
    ###################################################


    #WE PLOT TI
    plt.figure(figsize=(14, 7))
    
    plt.plot(epochData[0:,0], Ti275, color="orange")

    #print(epochData[0:,0])
    # plt.plot(Ti275)
    
    ''' 
    print('JUST EPOCHDATA')
    print(epochData)
    print('---------------------------------')
    print('ACTUAL TI')
    print(Ti275)
    ''' 

    plt.scatter(peaks_x[0:,0], peaks_y, color="blue")
    plt.scatter(plunges_x[0:,0], plunges_y, color="red")
    
    
    plt.title('Ti at 275 km_' + filename)
    plt.grid(True)
    plt.xlabel('Epoch Time')
    plt.ylabel('Line-of-Sight Ion Temperature [K]')
    plotfile = 'Andy PFISR/275{} - {}.png'.format(filename.split("_")[0], sys.argv[1])
    plt.savefig(plotfile)
    plt.cla()   # Clear axis
    plt.clf()   # Clear figure
    plt.close() # Close a figure window
    #plt.show()
         